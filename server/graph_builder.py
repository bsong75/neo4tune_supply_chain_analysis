"""Ingest EV battery supply chain data into Neo4j.

Entities become nodes with labels based on their type (Mine, Refinery,
ComponentMfg, CellMfg, PackAssembly, OEM). Links become SUPPLIES_TO
relationships between entities.
"""

import logging
import csv
import json
import io
from datetime import datetime, timezone

logger = logging.getLogger(__name__)

VALID_ENTITY_TYPES = {'Mine', 'Refinery', 'ComponentMfg', 'CellMfg', 'PackAssembly', 'OEM'}

TARGET_ENTITY_FIELDS = [
    'entity_id', 'type', 'name', 'country', 'material', 'component',
    'capacity', 'capacity_unit', 'tier', 'status',
]

TARGET_LINK_FIELDS = [
    'source_id', 'target_id', 'material', 'volume',
    'lead_time_days', 'cost_per_unit',
]

_indexes_created = False


def ensure_indexes(driver):
    """Create indexes and constraints for supply chain entities."""
    global _indexes_created
    if _indexes_created:
        return

    with driver.session() as session:
        for label in VALID_ENTITY_TYPES:
            session.run(
                f"CREATE CONSTRAINT IF NOT EXISTS FOR (e:{label}) REQUIRE e.entity_id IS UNIQUE"
            ).consume()
        session.run("CREATE INDEX IF NOT EXISTS FOR (rz:RiskZone) ON (rz.zone_id)").consume()

    _indexes_created = True
    logger.info("Neo4j indexes ensured")


def ingest_entities(driver, records):
    """Ingest a list of entity dicts into Neo4j.

    Each entity becomes a node with a label matching its type field.
    Returns: {entities_created}
    """
    ensure_indexes(driver)
    now = datetime.now(timezone.utc).isoformat()
    entities_created = 0

    with driver.session() as session:
        for rec in records:
            entity_type = str(rec.get('type', '')).strip()
            if entity_type not in VALID_ENTITY_TYPES:
                logger.warning("Skipping unknown entity type: %s", entity_type)
                continue

            entity_id = str(rec.get('entity_id', '')).strip()
            if not entity_id:
                continue

            session.run(f"""
                MERGE (e:{entity_type} {{entity_id: $entity_id}})
                ON CREATE SET
                    e.name = $name,
                    e.country = $country,
                    e.material = $material,
                    e.component = $component,
                    e.capacity = $capacity,
                    e.capacity_unit = $capacity_unit,
                    e.tier = $tier,
                    e.status = $status,
                    e.created_at = $now
            """,
                entity_id=entity_id,
                name=rec.get('name', '').strip(),
                country=rec.get('country', '').strip(),
                material=rec.get('material', '').strip(),
                component=rec.get('component', '').strip(),
                capacity=rec.get('capacity', '').strip(),
                capacity_unit=rec.get('capacity_unit', '').strip(),
                tier=rec.get('tier', '').strip(),
                status=rec.get('status', 'active').strip(),
                now=now,
            ).consume()
            entities_created += 1

    logger.info("Ingested %d entities", entities_created)
    return {'entities_created': entities_created}


def ingest_links(driver, records):
    """Ingest a list of link dicts into Neo4j as SUPPLIES_TO relationships.

    Returns: {links_created}
    """
    links_created = 0

    with driver.session() as session:
        for rec in records:
            source_id = str(rec.get('source_id', '')).strip()
            target_id = str(rec.get('target_id', '')).strip()
            if not source_id or not target_id:
                continue

            vol_str = str(rec.get('volume', '0')).strip()
            try:
                volume = float(vol_str)
            except ValueError:
                volume = 0.0

            lt_str = str(rec.get('lead_time_days', '0')).strip()
            try:
                lead_time = float(lt_str)
            except ValueError:
                lead_time = 0.0

            cost_str = str(rec.get('cost_per_unit', '0')).strip()
            try:
                cost = float(cost_str)
            except ValueError:
                cost = 0.0

            result = session.run("""
                MATCH (a {entity_id: $source_id})
                MATCH (b {entity_id: $target_id})
                MERGE (a)-[r:SUPPLIES_TO]->(b)
                ON CREATE SET
                    r.material = $material,
                    r.volume = $volume,
                    r.lead_time_days = $lead_time,
                    r.cost_per_unit = $cost
                RETURN count(r) AS cnt
            """,
                source_id=source_id,
                target_id=target_id,
                material=rec.get('material', '').strip(),
                volume=volume,
                lead_time=lead_time,
                cost=cost,
            )
            cnt = result.single()['cnt']
            links_created += cnt

    logger.info("Ingested %d links", links_created)
    return {'links_created': links_created}


def parse_csv(file_content):
    """Parse CSV string into list of dicts."""
    reader = csv.DictReader(io.StringIO(file_content))
    return [row for row in reader]


def parse_json(file_content):
    """Parse JSON string into list of dicts."""
    data = json.loads(file_content)
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and 'records' in data:
        return data['records']
    return [data]


def parse_xlsx(file_bytes):
    """Parse Excel .xlsx bytes into list of dicts."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"column_{i}" for i, h in enumerate(rows[0])]
    records = []
    for row in rows[1:]:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = str(val).strip() if val is not None else ''
        records.append(record)
    return records


def extract_preview(records, max_rows=5):
    """Return column names and up to max_rows of sample data."""
    if not records:
        return [], []
    columns = list(records[0].keys())
    preview = records[:max_rows]
    return columns, preview


def clear_all(driver):
    """Delete all nodes and relationships."""
    with driver.session() as session:
        session.run("MATCH (n) DETACH DELETE n").consume()
    logger.info("Database cleared")
